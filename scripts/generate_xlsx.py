import os
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/home/romankurmash/projects/Diploma/documentation/diploma/xlsx"
os.makedirs(OUT_DIR, exist_ok=True)

def create_styled_sheet(filename, sheet_title, headers, data, center_cols=None, left_cols=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.views.sheetView[0].showGridLines = True

    font_family = "Arial"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, bold=False)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        
    for row_idx, row_data in enumerate(data, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
                
            if center_cols and col_idx in center_cols:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    ws.row_dimensions[1].height = 28
    for r in range(2, len(data) + 2):
        ws.row_dimensions[r].height = 22

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out_path = os.path.join(OUT_DIR, filename)
    wb.save(out_path)
    print(f"Saved: {out_path}")

# Table 1: Consultants
create_styled_sheet(
    "consultants.xlsx", "Консультанти",
    ["Розділ", "Прізвище, ініціали консультанта", "Завдання видав", "Завдання прийняв"],
    [["1-4", "Павловський Т. М.", "", ""]],
    center_cols=[1, 3, 4]
)

# Table 2: Calendar Plan
create_styled_sheet(
    "calendar_plan.xlsx", "Календарний план",
    ["№ з/п", "Назва етапів дипломної роботи", "Термін виконання", "Примітка"],
    [
        [1, "Вибір та затвердження теми дипломної роботи", "Вересень 2025", ""],
        [2, "Збір та аналіз літературних джерел", "Жовтень 2025", ""],
        [3, "Аналіз предметної області, розробка ТЗ", "Листопад 2025", ""],
        [4, "Проектування архітектури системи", "Грудень 2025", ""],
        [5, "Розробка серверної частини (Flask API)", "Січень 2026", ""],
        [6, "Розробка клієнтської частини (SPA Dashboard)", "Лютий 2026", ""],
        [7, "Інтеграція з LLM та системою моніторингу", "Березень 2026", ""],
        [8, "Реалізація модуля аналізу PCAP", "Квітень 2026", ""],
        [9, "Тестування та налагодження", "Квітень 2026", ""],
        [10, "Оформлення пояснювальної записки", "Травень 2026", ""],
        [11, "Підготовка до захисту", "Червень 2026", ""]
    ],
    center_cols=[1, 3]
)

# Table 3: Information Assets (Table 1.1)
create_styled_sheet(
    "info_assets.xlsx", "Інформаційні активи",
    ["Актив", "Критичність", "Тип даних", "Загрози"],
    [
        ["Веб-сервер Nginx", "Висока", "HTTP-трафік, заголовки", "DDoS, SQL Injection, XSS"],
        ["WordPress", "Висока", "Контент, облікові дані", "Brute-force, RCE, Unauthorized Access"],
        ["MySQL БД", "Критична", "Дані користувачів", "SQL Injection, Data Exfiltration"],
        ["Логи системи", "Середня", "Системні журнали", "Tampering, Information Disclosure"],
        ["Мережевий трафік", "Висока", "Пакети даних", "Sniffing, Man-in-the-Middle"]
    ],
    center_cols=[2]
)

# Table 4: Solutions Comparison (Table 1.2)
create_styled_sheet(
    "solutions_comparison.xlsx", "Порівняльний аналіз",
    ["Критерій", "Snort", "Suricata", "Darktrace", "ELK Stack", "NetVigil (наша)"],
    [
        ["Вартість", "Безкоштовно", "Безкоштовно", "Комерційна", "Безкоштовно", "Безкоштовно"],
        ["AI-аналіз", "Ні", "Ні", "Так", "Ні", "Так (LLM)"],
        ["Zero-day виявлення", "Ні", "Ні", "Так", "Ні", "Так"],
        ["Веб-інтерфейс", "Обмежений", "Обмежений", "Так", "Так", "Так"],
        ["Пояснення загроз", "Ні", "Ні", "Частково", "Ні", "Так (LLM)"],
        ["PCAP-аналіз", "Так", "Так", "Ні", "Ні", "Так"],
        ["Чат з AI", "Ні", "Ні", "Ні", "Ні", "Так"],
        ["Ресурсоємність", "Низька", "Середня", "Висока", "Висока", "Середня"]
    ]
)

# Table 5: STRIDE Threats (Table 2.1)
create_styled_sheet(
    "stride_threats.xlsx", "Модель STRIDE",
    ["Категорія", "Загроза", "Компонент", "Ймовірність", "Вплив", "Заходи захисту"],
    [
        ["S — Spoofing", "Підміна IP-адреси відправника", "Nginx Proxy", "Середня", "Високий", "Аналіз паттернів LLM, логування"],
        ["T — Tampering", "Модифікація HTTP-запитів", "WordPress", "Висока", "Критичний", "WAF-правила, LLM-аналіз payloads"],
        ["R — Repudiation", "Відмова від дій у системі", "Всі сервіси", "Низька", "Середній", "Централізоване логування (Loki)"],
        ["I — Information Disclosure", "Витік даних через SQL-ін'єкції", "MySQL", "Висока", "Критичний", "LLM-аналіз SQL-паттернів"],
        ["D — Denial of Service", "DDoS через надмірні HTTP-запити", "Nginx", "Середня", "Високий", "Rate limiting, моніторинг метрик"],
        ["E — Elevation of Privilege", "Отримання доступу admin через brute-force", "WordPress", "Висока", "Критичний", "Виявлення brute-force через LLM"]
    ],
    center_cols=[4, 5]
)

# Table 6: Technology Stack (Table 3.1)
create_styled_sheet(
    "tech_stack.xlsx", "Технологічний стек",
    ["Компонент", "Технологія", "Версія", "Призначення"],
    [
        ["Backend", "Python", "3.11", "Мова програмування серверної частини"],
        ["Web Framework", "Flask", "3.1.0", "REST API фреймворк"],
        ["WSGI Server", "Gunicorn", "23.0.0", "Production-сервер"],
        ["HTTP Client", "httpx", "0.28.0", "Асинхронні HTTP-запити до Loki/Ollama"],
        ["Frontend", "HTML5 + CSS3 + JS", "ES6+", "Клієнтська частина SPA"],
        ["Charts", "Chart.js", "4.4.4", "Побудова графіків"],
        ["Fonts", "Inter, JetBrains Mono", "—", "Типографіка"],
        ["LLM", "Llama-3 8B", "—", "Велика мовна модель"],
        ["LLM Platform", "Ollama", "Latest", "Self-hosted LLM-платформа"],
        ["Log Storage", "Grafana Loki", "Latest", "Зберігання логів"],
        ["Log Agent", "Promtail", "Latest", "Збір логів контейнерів"],
        ["Metrics", "Prometheus", "Latest", "Збір метрик"],
        ["Alerting", "Alertmanager", "Latest", "Управління алертами"],
        ["Containerization", "Docker", "24.x", "Контейнеризація"],
        ["Orchestration", "Docker Compose", "2.x", "Оркестрація контейнерів"],
        ["CI/CD", "Jenkins", "LTS", "Автоматизація розгортання"]
    ],
    center_cols=[3]
)

# Table 7: API Endpoints (Table 3.2)
create_styled_sheet(
    "api_endpoints.xlsx", "API Endpoints",
    ["Метод", "Endpoint", "Опис"],
    [
        ["GET", "/api/dashboard", "Агреговані KPI-дані для дашборду"],
        ["GET", "/api/dashboard/timeline", "Дані для графіку timeline за 6 годин"],
        ["GET", "/api/logs", "Отримання логів з фільтрацією"],
        ["POST", "/api/analyze", "Аналіз текстових логів через LLM"],
        ["POST", "/api/analyze/pcap", "Завантаження та аналіз PCAP-файлу"],
        ["GET", "/api/alerts", "Список алертів з Alertmanager"],
        ["POST", "/api/chat", "Чат з AI-асистентом"],
        ["GET", "/api/metrics", "Метрики системи з Prometheus"],
        ["GET", "/api/health", "Health-check сервісу"]
    ],
    center_cols=[1]
)

# Table 8: API Testing (Table 4.1)
create_styled_sheet(
    "api_testing.xlsx", "Тестування API",
    ["Endpoint", "Метод", "Тест-кейс", "Очікуваний результат", "Фактичний результат", "Статус"],
    [
        ["/api/health", "GET", "Перевірка доступності", "JSON {status: ok}", "JSON {status: ok}", "Pass"],
        ["/api/dashboard", "GET", "Отримання KPI", "JSON з метриками", "JSON з метриками", "Pass"],
        ["/api/logs", "GET", "Фільтр за контейнером", "Логи nginx-proxy", "Логи nginx-proxy", "Pass"],
        ["/api/logs?keyword=error", "GET", "Пошук за словом", "Логи з 'error'", "Логи з 'error'", "Pass"],
        ["/api/analyze", "POST", "Аналіз логів LLM", "JSON з risk_score", "JSON з risk_score", "Pass"],
        ["/api/analyze/pcap", "POST", "Завантаження PCAP", "Аналіз пакетів", "Аналіз пакетів", "Pass"],
        ["/api/analyze/pcap", "POST", "Файл > 10MB", "Помилка 400", "Помилка 400", "Pass"],
        ["/api/analyze/pcap", "POST", "Невірний формат", "Помилка 400", "Помилка 400", "Pass"],
        ["/api/alerts", "GET", "Отримання алертів", "JSON масив", "JSON масив", "Pass"],
        ["/api/chat", "POST", "Відправка повідомлення", "Відповідь LLM", "Відповідь LLM", "Pass"],
        ["/api/chat", "POST", "Порожнє повідомлення", "Помилка 400", "Помилка 400", "Pass"],
        ["/api/metrics", "GET", "Отримання метрик", "JSON з CPU, RAM", "JSON з CPU, RAM", "Pass"]
    ],
    center_cols=[2, 6]
)

# Table 9: Threat Detection Testing (Table 4.2)
create_styled_sheet(
    "threat_detection_testing.xlsx", "Тестування загроз",
    ["Тип атаки", "Кількість тестів", "Виявлено", "Невиявлено", "Точність"],
    [
        ["SQL Injection", 10, 9, 1, "90%"],
        ["Brute-force", 10, 10, 0, "100%"],
        ["XSS", 10, 8, 2, "80%"],
        ["Port Scan (PCAP)", 5, 5, 0, "100%"],
        ["DDoS Pattern", 5, 4, 1, "80%"],
        ["Загалом", 40, 36, 4, "90%"]
    ],
    center_cols=[2, 3, 4, 5]
)

# Table 10: Load Testing (Table 4.3)
create_styled_sheet(
    "load_testing.xlsx", "Навантажувальне тестування",
    ["Endpoint", "Потоків", "З'єднань", "Тривалість", "RPS", "Avg Latency", "Max Latency"],
    [
        ["/api/dashboard", 4, 50, "30s", 125, "380ms", "1.2s"],
        ["/api/logs", 4, 50, "30s", 95, "520ms", "1.8s"],
        ["/api/metrics", 4, 50, "30s", 210, "230ms", "0.9s"],
        ["/api/health", 4, 50, "30s", 850, "58ms", "0.3s"],
        ["/api/analyze (LLM)", 4, 10, "30s", 0.8, "12s", "45s"]
    ],
    center_cols=[2, 3, 4, 5, 6, 7]
)

# Table 11: Abbreviations Table
create_styled_sheet(
    "abbreviations.xlsx", "Скорочення",
    ["Скорочення", "Розшифрування"],
    [
        ["API", "Application Programming Interface — інтерфейс програмування додатків"],
        ["CI/CD", "Continuous Integration / Continuous Deployment — безперервна інтеграція та розгортання"],
        ["CSS", "Cascading Style Sheets — каскадні таблиці стилів"],
        ["DDoS", "Distributed Denial of Service — розподілена відмова в обслуговуванні"],
        ["DFD", "Data Flow Diagram — діаграма потоку даних"],
        ["DNS", "Domain Name System — система доменних імен"],
        ["GPU", "Graphics Processing Unit — графічний процесор"],
        ["HTML", "HyperText Markup Language — мова розмітки гіпертексту"],
        ["HTTP/HTTPS", "HyperText Transfer Protocol (Secure) — протокол передачі гіпертексту"],
        ["IDS", "Intrusion Detection System — система виявлення вторгнень"],
        ["IPS", "Intrusion Prevention System — система запобігання вторгненням"],
        ["JSON", "JavaScript Object Notation — формат обміну даними"],
        ["LLM", "Large Language Model — велика мовна модель"],
        ["PCAP", "Packet Capture — формат захоплення мережевих пакетів"],
        ["REST", "Representational State Transfer — архітектурний стиль API"],
        ["SIEM", "Security Information and Event Management — управління інформацією та подіями безпеки"],
        ["SPA", "Single Page Application — односторінковий додаток"],
        ["SQL", "Structured Query Language — структурована мова запитів"],
        ["SQLi", "SQL Injection — впровадження SQL-коду"],
        ["SSH", "Secure Shell — протокол безпечного доступу"],
        ["SSL/TLS", "Secure Sockets Layer / Transport Layer Security — протоколи шифрування"],
        ["STRIDE", "Spoofing, Tampering, Repudiation, Information disclosure, Denial of service, Elevation of privilege — модель загроз"],
        ["TCP", "Transmission Control Protocol — протокол керування передачею"],
        ["UDP", "User Datagram Protocol — протокол датаграм користувача"],
        ["UML", "Unified Modeling Language — уніфікована мова моделювання"],
        ["URL", "Uniform Resource Locator — уніфікований покажчик ресурсу"],
        ["VPN", "Virtual Private Network — віртуальна приватна мережа"],
        ["WAF", "Web Application Firewall — міжмережевий екран веб-додатків"],
        ["XSS", "Cross-Site Scripting — міжсайтовий скриптинг"],
        ["БД", "База даних"],
        ["ІБ", "Інформаційна безпека"],
        ["ПЗ", "Програмне забезпечення"],
        ["ТЗ", "Технічне завдання"]
    ],
    center_cols=[1]
)

# Table 12: Appendix A Functional Requirements (4.1)
create_styled_sheet(
    "functional_requirements.xlsx", "Функціональні вимоги",
    ["№", "Функція", "Опис"],
    [
        [1, "Dashboard", "Інтерактивна панель з KPI (події, помилки, алерти, CPU), графіками Timeline та метриками системи"],
        [2, "Log Viewer", "Перегляд логів з фільтрацією за контейнером, часом, ключовими словами. Ручний AI-аналіз"],
        [3, "Traffic Analysis", "Завантаження PCAP-файлів, парсинг пакетів, AI-аналіз загроз"],
        [4, "Alerts", "Перегляд алертів безпеки з Alertmanager"],
        [5, "AI Chat", "Чат-інтерфейс для діалогу з LLM"],
        [6, "REST API", "9 endpoints для взаємодії фронтенду з бекендом"],
        [7, "LLM Integration", "Класифікація за 10 категоріями загроз, risk_score 0-10"],
        [8, "Auto-alerting", "Автоматичне сповіщення через Alertmanager/Telegram"]
    ],
    center_cols=[1]
)

# Table 13: Appendix A Development Stages (7)
create_styled_sheet(
    "development_stages.xlsx", "Етапи розробки",
    ["Етап", "Зміст", "Термін"],
    [
        [1, "Аналіз предметної області", "Вересень — Листопад 2025"],
        [2, "Проектування архітектури", "Грудень 2025"],
        [3, "Розробка серверної частини", "Січень — Лютий 2026"],
        [4, "Розробка клієнтської частини", "Лютий — Березень 2026"],
        [5, "Інтеграція з LLM та PCAP", "Березень — Квітень 2026"],
        [6, "Тестування та налагодження", "Квітень — Травень 2026"],
        [7, "Оформлення документації", "Травень — Червень 2026"]
    ],
    center_cols=[1, 3]
)
